# tests/aliases/test_openpyxl.py
# C-020: Arabic aliases for openpyxl

import pathlib

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed")

ALIASES_DIR = pathlib.Path(__file__).parent.parent.parent / "arabicpython" / "aliases"


@pytest.fixture(scope="module")
def جداول_اكسل():
    """Return a ModuleProxy wrapping `openpyxl`."""
    from arabicpython.aliases._finder import AliasFinder

    finder = AliasFinder(mappings_dir=ALIASES_DIR)
    spec = finder.find_spec("جداول_اكسل", None, None)
    assert spec is not None, "AliasFinder did not find 'جداول_اكسل'"
    proxy = spec.loader.create_module(spec)
    assert proxy is not None
    return proxy


class TestOpenpyxlCore:
    def test_workbook_and_sheet_classes(self, جداول_اكسل):
        import openpyxl.cell.cell
        import openpyxl.worksheet.worksheet

        from arabicpython.aliases._proxy import ClassFactory

        assert isinstance(جداول_اكسل.مصنف, ClassFactory)
        assert جداول_اكسل.ورقه_عمل is openpyxl.worksheet.worksheet.Worksheet
        assert جداول_اكسل.خليه is openpyxl.cell.cell.Cell
        assert جداول_اكسل.افتح_مصنف is openpyxl.load_workbook

    def test_style_and_chart_aliases(self, جداول_اكسل):
        import openpyxl.chart
        import openpyxl.styles

        assert جداول_اكسل.خط is openpyxl.styles.Font
        assert جداول_اكسل.تعبئه_نمط is openpyxl.styles.PatternFill
        assert جداول_اكسل.مخطط_اعمده is openpyxl.chart.BarChart
        assert جداول_اكسل.مرجع_رسم is openpyxl.chart.Reference


class TestOpenpyxlFunctional:
    def test_workbook_instance_methods_and_save(self, جداول_اكسل, tmp_path):
        workbook = جداول_اكسل.مصنف()
        wrapped = object.__getattribute__(workbook, "_wrapped")

        assert workbook.ورقه_نشطه is wrapped.active
        assert workbook.انشئ_ورقه == wrapped.create_sheet

        sheet = workbook.ورقه_نشطه
        sheet.title = "Scores"
        sheet.append(["name", "score"])
        sheet.append(["Layla", 95])

        output = tmp_path / "scores.xlsx"
        workbook.احفظ_مصنف(output)

        assert output.exists()
        assert جداول_اكسل.افتح_مصنف(output).active["B2"].value == 95

    def test_worksheet_and_cell_attributes_work_with_class_proxy(self):
        from arabicpython.aliases import ClassProxy, load_mapping

        mapping = load_mapping(ALIASES_DIR / "openpyxl.toml")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet_proxy = ClassProxy(sheet, mapping.attributes)

        sheet_proxy.عنوان = "Data"
        sheet_proxy.اضف_صف(["item", "value"])
        cell = sheet_proxy.خليه(row=1, column=2)
        cell_proxy = ClassProxy(cell, mapping.attributes)
        cell_proxy.قيمه = "score"

        assert sheet.title == "Data"
        assert sheet.max_row == 1
        assert cell.value == "score"


class TestOpenpyxlTomlMeta:
    def test_toml_parseable(self):
        import tomllib

        p = ALIASES_DIR / "openpyxl.toml"
        with open(p, "rb") as f:
            data = tomllib.load(f)
        assert data["meta"]["python_module"] == "openpyxl"
        assert data["meta"]["arabic_name"] == "جداول_اكسل"

    def test_entry_count(self):
        import tomllib

        p = ALIASES_DIR / "openpyxl.toml"
        with open(p, "rb") as f:
            data = tomllib.load(f)
        assert len(data["entries"]) >= 25
